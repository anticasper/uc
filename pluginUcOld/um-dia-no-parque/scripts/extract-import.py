#!/usr/bin/env python3
"""
Extract UC data + Atividades from XLSX form responses → JSON for WP-CLI import.

Usage:
  python3 scripts/extract-import.py <xlsx_path> <output_json>
"""
import json, sys, os
from collections import OrderedDict
from openpyxl import load_workbook

XLSX = sys.argv[1]
OUT  = sys.argv[2] if len(sys.argv) > 2 else 'import-data.json'

wb = load_workbook(XLSX, data_only=True)
ws = wb['Respostas ao formulário 1']

def cell(v):
    if v is None: return ''
    s = str(v).strip()
    return s if s.lower() not in ('none', 'n/a', '') else ''

def norm(name):
    import unicodedata
    n = cell(name).lower()
    n = unicodedata.normalize('NFKD', n).encode('ascii', 'ignore').decode()
    return ' '.join(n.split())

ucs = []
atividades = []
total_rows = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    total_rows += 1
    vals = list(row)

    uc_name = cell(vals[3] if len(vals) > 3 else '')
    if not uc_name:
        continue

    # UC data (columns C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, L=12, Q=17, R=18)
    uc_entry = OrderedDict([
        ('uc_name', uc_name),
        ('uc_name_norm', norm(uc_name)),
        ('responsavel', cell(vals[4] if len(vals) > 4 else '')),
        ('whatsapp', cell(vals[5] if len(vals) > 5 else '')),
        ('orgao', cell(vals[6] if len(vals) > 6 else '')),
        ('bioma', cell(vals[7] if len(vals) > 7 else '')),
        ('estado', cell(vals[8] if len(vals) > 8 else '')),
        ('municipio', cell(vals[9] if len(vals) > 9 else '')),
        ('telefone', cell(vals[10] if len(vals) > 10 else '')),
        ('redes_sociais', cell(vals[12] if len(vals) > 12 else '')),
        ('cep', cell(vals[17] if len(vals) > 17 else '')),
        ('endereco', cell(vals[18] if len(vals) > 18 else '')),
    ])
    ucs.append(uc_entry)

    # Atividades (columns S=19, T=20, U=21, V=22)
    slots = [
        {'desc_col': 19, 'date_col': 25, 'time_col': 31, 'title_col': 32,
         'short_col': 33, 'audience_col': 34, 'diff_col': 35, 'bring_col': 36, 'contact_col': 37},
        {'desc_col': 20, 'date_col': 38, 'time_col': 39, 'title_col': 40,
         'short_col': 41, 'audience_col': 42, 'diff_col': 43, 'bring_col': 44, 'contact_col': 0},
        {'desc_col': 21, 'date_col': 46, 'time_col': 47, 'title_col': 48,
         'short_col': 49, 'audience_col': 50, 'diff_col': 51, 'bring_col': 52, 'contact_col': 0},
        {'desc_col': 22, 'date_col': 53, 'time_col': 54, 'title_col': 55,
         'short_col': 56, 'audience_col': 57, 'diff_col': 58, 'bring_col': 59, 'contact_col': 0},
    ]

    for slot_idx, s in enumerate(slots, 1):
        desc = cell(vals[s['desc_col'] - 1] if len(vals) > s['desc_col'] - 1 else '')
        if not desc:
            continue

        title = cell(vals[s['title_col'] - 1] if s['title_col'] and len(vals) > s['title_col'] - 1 else '')
        if not title:
            title = desc[:60].rstrip('.,;: ') + ('...' if len(desc) > 60 else '')

        date   = cell(vals[s['date_col'] - 1] if s['date_col'] and len(vals) > s['date_col'] - 1 else '')
        time   = cell(vals[s['time_col'] - 1] if s['time_col'] and len(vals) > s['time_col'] - 1 else '')
        short  = cell(vals[s['short_col'] - 1] if s['short_col'] and len(vals) > s['short_col'] - 1 else '')
        audience = cell(vals[s['audience_col'] - 1] if s['audience_col'] and len(vals) > s['audience_col'] - 1 else '')
        diff   = cell(vals[s['diff_col'] - 1] if s['diff_col'] and len(vals) > s['diff_col'] - 1 else '')
        bring  = cell(vals[s['bring_col'] - 1] if s['bring_col'] and len(vals) > s['bring_col'] - 1 else '')
        contact = cell(vals[s['contact_col'] - 1] if s['contact_col'] and len(vals) > s['contact_col'] - 1 else '')

        # Clean date
        if date and len(date) > 10 and date[4] == '-':
            date = date[:10]

        # Normalize difficulty
        diff_map = {'leve':'facil','fácil':'facil','facil':'facil',
                    'moderado':'moderado','moderada':'moderado',
                    'difícil':'dificil','dificil':'dificil'}
        diff_key = diff_map.get(diff.lower().strip(), '')

        atividade = OrderedDict([
            ('uc_name', uc_name),
            ('slot', slot_idx),
            ('title', title),
            ('description', desc),
            ('short_description', short),
            ('audience', audience),
            ('difficulty', diff_key),
            ('difficulty_raw', diff),
            ('date', date),
            ('time', time),
            ('what_to_bring', bring),
            ('contact', contact),
        ])
        atividades.append(atividade)

output = {
    'total_rows': total_rows,
    'total_ucs': len(ucs),
    'total_atividades': len(atividades),
    'ucs': ucs,
    'atividades': atividades,
}

os.makedirs(os.path.dirname(OUT) if os.path.dirname(OUT) else '.', exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Extraído: {len(ucs)} UCs, {len(atividades)} atividades → {OUT}")
